"""
Roster parsing + the anomaly agent.

Everything here is pure (no Flask, DB, network or disk state) so it can be
unit-tested offline, mirroring how the AI-designer agent is tested with a fake
LLM. `rapidfuzz` (already a dependency) powers near-duplicate detection.

Flow:
    rows            = parse_table(filename, file_bytes)     # list[dict] keyed by header
    mapped, unmapped= map_rows(rows, placeholders, mapping) # header -> placeholder
    report          = detect_anomalies(mapped, name_field)  # findings + fixes
    cleaned         = suggest_rows(mapped, report)           # safe fixes pre-applied
"""

import io
import csv
import re

from rapidfuzz import fuzz

# A near-duplicate is flagged when two distinct names score at/above this.
NEAR_DUP_THRESHOLD = 88

# Upload guards
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_ROWS = 2000

# Common roster header wordings -> template placeholder. Matched on the
# normalized key, so "Full Name", "full_name" and "FULLNAME" all hit the same entry.
HEADER_ALIASES = {
    "RECIPIENT_NAME": [
        "name", "fullname", "full name", "candidatename", "participantname",
        "studentname", "recipient", "recipientname", "attendee", "attendeename",
        "employeename", "person", "awardee", "learner", "membername",
    ],
    "EVENT_NAME": [
        "event", "eventname", "course", "coursename", "program", "programme",
        "programname", "workshop", "training", "reason", "achievement", "activity",
    ],
    "ISSUE_DATE": [
        "date", "issuedate", "issuedon", "dateofissue", "awardeddate",
        "completiondate", "dateissued", "eventdate",
    ],
    "SIGNATORY_NAME": ["signatory", "signedby", "authorisedby", "authorizedby", "signature"],
    "SIGNATORY_TITLE": ["designation", "title", "signatorytitle", "role", "position"],
    "ORG_NAME": ["organisation", "organization", "org", "company", "institute", "institution", "college"],
}


class ParseError(Exception):
    """Raised when an uploaded file cannot be read as a table."""


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

def parse_table(filename, file_bytes):
    """
    Parse an uploaded roster into a list of ordered dicts (header -> value).

    Supports .csv/.tsv (stdlib) and .xlsx (openpyxl, first sheet, row 1 = header).
    Blank rows are skipped. Raises ParseError on an unreadable or headerless file.
    """
    if not file_bytes:
        raise ParseError("The uploaded file is empty.")
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ParseError(
            "File is larger than {} MB. Split the roster and upload in parts.".format(
                MAX_UPLOAD_BYTES // (1024 * 1024)
            )
        )

    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(file_bytes)
    if name.endswith((".csv", ".tsv", ".txt")) or not name:
        # Delimiter is sniffed, not assumed: Excel exports in many locales use
        # ';', and '.csv' files are frequently tab- or pipe-separated.
        return _parse_delimited(file_bytes)
    raise ParseError(
        "Unsupported file type '{}'. Upload a .csv, .tsv or .xlsx file.".format(name)
    )


def _sniff_delimiter(text):
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        # Sniffer fails on single-column files; fall back to the most frequent candidate.
        counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","


def _find_header_row(rows_raw):
    """
    Locate the header row. Spreadsheets exported from event tools often carry a
    title/blank/notes preamble, so the first non-empty row is not always the
    header: pick the first row with >= 2 non-empty cells, else the first
    non-empty row.
    """
    first_non_empty = None
    for i, row in enumerate(rows_raw):
        filled = [c for c in row if (c or "").strip()]
        if not filled:
            continue
        if first_non_empty is None:
            first_non_empty = i
        if len(filled) >= 2:
            return i
    if first_non_empty is None:
        raise ParseError("The file has no readable rows.")
    return first_non_empty


def _dedupe_headers(headers):
    """'Name', 'Name' -> 'Name', 'Name_2' so later columns aren't silently lost."""
    seen, out = {}, []
    for h in headers:
        h = (h or "").strip()
        if not h:
            out.append("")
            continue
        seen[h] = seen.get(h, 0) + 1
        out.append(h if seen[h] == 1 else "{}_{}".format(h, seen[h]))
    return out


def _decode(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ParseError("Could not decode the file — save it as UTF-8 and retry.")


def _parse_delimited(file_bytes, delimiter=None):
    text = _decode(file_bytes)
    delimiter = delimiter or _sniff_delimiter(text)
    try:
        rows_raw = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    except csv.Error as e:
        raise ParseError("Could not parse the file as a table: {}".format(e))
    if not rows_raw:
        raise ParseError("The file is empty.")

    header_idx = _find_header_row(rows_raw)
    headers = _dedupe_headers(rows_raw[header_idx])
    if not any(headers):
        raise ParseError("Could not find a header row with column names.")

    rows = []
    for raw in rows_raw[header_idx + 1:]:
        if not any((c or "").strip() for c in raw):
            continue  # skip fully blank rows
        row = {}
        for i, header in enumerate(headers):
            if header:
                row[header] = (raw[i] if i < len(raw) else "").strip()
        rows.append(row)
    return rows


def _parse_xlsx(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ParseError(
            "Excel support needs openpyxl (`pip install openpyxl`). "
            "Alternatively upload the sheet as .csv."
        )
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - surface a clean message
        raise ParseError("Could not read the Excel file: {}".format(e))

    ws = wb.active
    all_rows = [
        ["" if c is None else str(c).strip() for c in raw]
        for raw in ws.iter_rows(values_only=True)
    ]
    if not all_rows:
        raise ParseError("The spreadsheet is empty.")

    header_idx = _find_header_row(all_rows)
    headers = _dedupe_headers(all_rows[header_idx])
    if not any(headers):
        raise ParseError("Could not find a header row with column names.")

    rows = []
    for raw in all_rows[header_idx + 1:]:
        cells = raw
        if not any(cells):
            continue
        row = {}
        for i, header in enumerate(headers):
            if header:
                row[header] = cells[i] if i < len(cells) else ""
        rows.append(row)
    return rows


# --------------------------------------------------------------------------- #
# Column mapping (roster header -> template placeholder)
# --------------------------------------------------------------------------- #

def _norm_key(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


_ALIAS_LOOKUP = {
    _norm_key(alias): placeholder
    for placeholder, aliases in HEADER_ALIASES.items()
    for alias in aliases
}

# Below this fuzzy score a header is left unmapped rather than guessed wrongly.
HEADER_MATCH_THRESHOLD = 82


def match_header(header, placeholders):
    """
    Resolve one roster header to a placeholder.
    Returns (placeholder|None, confidence 0-100, how) where `how` is
    'exact' | 'alias' | 'fuzzy' | 'none' — surfaced in the UI so the user can
    see (and override) what the importer decided.
    """
    key = _norm_key(header)
    if not key:
        return None, 0, "none"

    norm_to_ph = {_norm_key(p): p for p in placeholders}
    if key in norm_to_ph:
        return norm_to_ph[key], 100, "exact"

    alias_hit = _ALIAS_LOOKUP.get(key)
    if alias_hit and alias_hit in placeholders:
        return alias_hit, 95, "alias"

    # Fuzzy: tolerate typos/extra words ("Candidate Full Name ", "Recipient-Name")
    best, best_score = None, 0
    for ph in placeholders:
        score = fuzz.token_set_ratio(key, _norm_key(ph))
        for alias in HEADER_ALIASES.get(ph, []):
            score = max(score, fuzz.token_set_ratio(key, _norm_key(alias)))
        if score > best_score:
            best, best_score = ph, score
    if best and best_score >= HEADER_MATCH_THRESHOLD:
        return best, int(best_score), "fuzzy"
    return None, int(best_score), "none"


def map_rows(rows, placeholders, mapping=None):
    """
    Re-key each row from roster headers to template placeholder names.

    `mapping` (optional) is an explicit {header: PLACEHOLDER} dict from the user
    and always wins. Remaining headers are resolved by exact match, then a
    synonym table, then fuzzy matching above a confidence floor.

    Returns (mapped_rows, unmapped_placeholders, mapping_report) where the report
    lists every source header with the placeholder it was mapped to, the
    confidence and the method used.
    """
    mapping = mapping or {}
    placeholders = list(placeholders)

    headers = []
    for row in rows:
        for h in row.keys():
            if h not in headers:
                headers.append(h)

    resolved, report, taken = {}, [], set()
    for header in headers:
        if header in mapping and mapping[header]:
            target, score, how = mapping[header], 100, "manual"
        else:
            target, score, how = match_header(header, placeholders)
        # One placeholder can only be fed by one column; keep the stronger match.
        if target and target in taken:
            target, how = None, "duplicate_column"
        if target:
            taken.add(target)
        resolved[header] = target
        report.append({
            "HEADER": header, "PLACEHOLDER": target,
            "CONFIDENCE": score, "METHOD": how,
        })

    mapped = []
    for row in rows:
        out = {}
        for header, value in row.items():
            target = resolved.get(header)
            if target:
                out[target] = value
        mapped.append(out)

    unmapped = [p for p in placeholders if p not in taken]
    return mapped, unmapped, report


# --------------------------------------------------------------------------- #
# The anomaly agent
# --------------------------------------------------------------------------- #

def _clean_ws(value):
    return re.sub(r"\s+", " ", (value or "").strip())


def _title_case(value):
    # Title-case each word but keep short connectors lower except when first.
    small = {"of", "the", "and", "for", "de", "van", "da"}
    words = _clean_ws(value).split(" ")
    out = []
    for i, w in enumerate(words):
        lw = w.lower()
        if i > 0 and lw in small:
            out.append(lw)
        else:
            out.append(lw[:1].upper() + lw[1:] if lw else lw)
    return " ".join(out)


def _finding(row, field, kind, severity, current, suggested, message, **extra):
    f = {
        "row": row,
        "field": field,
        "type": kind,
        "severity": severity,
        "current": current,
        "suggested": suggested,
        "message": message,
    }
    f.update(extra)
    return f


def detect_anomalies(rows, name_field, required_fields=None):
    """
    Inspect mapped rows and return a structured report of data-quality issues
    with proposed fixes, for a human to approve (maker-checker).

    Detects:
      * missing        — a required field is blank (needs a human; no auto-fix)
      * whitespace     — leading/trailing/double spaces (safe auto-fix: trim)
      * casing         — ALL CAPS / all lower name (auto-fix: Title Case)
      * duplicate_exact— identical name in an earlier row (needs a human)
      * duplicate_near — fuzzily similar name in an earlier row (needs review)

    `name_field` is the placeholder used as the certificate's unique name.
    """
    required = list(required_fields) if required_fields else [name_field]
    anomalies = []
    seen = {}  # normalized name -> first row index

    for i, row in enumerate(rows):
        # missing required fields
        for field in required:
            if not (row.get(field) or "").strip():
                anomalies.append(_finding(
                    i, field, "missing", "high",
                    row.get(field, ""), None,
                    "Required field '{}' is empty.".format(field),
                ))

        name = row.get(name_field, "") or ""
        if not name.strip():
            continue

        # whitespace hygiene on the name field
        cleaned = _clean_ws(name)
        if cleaned != name:
            anomalies.append(_finding(
                i, name_field, "whitespace", "low",
                name, cleaned,
                "Extra whitespace — trim to '{}'.".format(cleaned),
            ))

        # casing on the name field
        if cleaned and (cleaned.isupper() or cleaned.islower()):
            titled = _title_case(cleaned)
            if titled != cleaned:
                anomalies.append(_finding(
                    i, name_field, "casing", "medium",
                    cleaned, titled,
                    "Inconsistent casing — suggest '{}'.".format(titled),
                ))

        # duplicate detection against earlier rows
        key = _clean_ws(name).lower()
        if key in seen:
            anomalies.append(_finding(
                i, name_field, "duplicate_exact", "high",
                cleaned, None,
                "Duplicate of row {} — will overwrite unless renamed.".format(seen[key] + 1),
                duplicate_of=seen[key],
            ))
        else:
            near = _nearest(key, seen)
            if near is not None:
                other_idx, score = near
                anomalies.append(_finding(
                    i, name_field, "duplicate_near", "medium",
                    cleaned, None,
                    "Looks similar to row {} ('{}', {}% match) — possible typo.".format(
                        other_idx + 1, rows[other_idx].get(name_field, ""), int(score)
                    ),
                    duplicate_of=other_idx, score=int(score),
                ))
            seen[key] = i

    counts = {}
    for a in anomalies:
        counts[a["type"]] = counts.get(a["type"], 0) + 1

    return {
        "total_rows": len(rows),
        "anomaly_count": len(anomalies),
        "counts": counts,
        "anomalies": anomalies,
    }


def _nearest(key, seen):
    """Return (row_index, score) of the closest earlier name at/above threshold."""
    best_idx, best_score = None, 0
    for other_key, idx in seen.items():
        score = fuzz.ratio(key, other_key)
        if score >= NEAR_DUP_THRESHOLD and score > best_score:
            best_idx, best_score = idx, score
    return (best_idx, best_score) if best_idx is not None else None


def suggest_rows(rows, report):
    """
    Return a deep copy of `rows` with the SAFE auto-fixes (whitespace, casing)
    from the report applied, so the UI can preview a cleaned roster. Duplicates
    and missing values are left untouched — those need a human decision.
    """
    fixed = [dict(r) for r in rows]
    for a in report.get("anomalies", []):
        if a["type"] in ("whitespace", "casing") and a["suggested"] is not None:
            fixed[a["row"]][a["field"]] = a["suggested"]
    return fixed
